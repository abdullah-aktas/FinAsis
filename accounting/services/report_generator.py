# -*- coding: utf-8 -*-
"""
Muhasebe Raporlama Sistemi - TFRS/VUK/TDHP Uyumlu
Türkiye Muhasebe Standartları, Vergi Usul Kanunu ve Türkiye Tekdüzen Hesap Planı'na uygun rapor üretimi
"""
import logging
from decimal import Decimal
from datetime import date
from typing import Dict, List, Optional, Any
from django.db.models import Sum
from django.utils import timezone
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from django.http import HttpResponse
import io
import hashlib

from ..models import (
    Company,
    GLAccount,
    GLJournalEntry,
    GLJournalLine,
)
from finance.accounting.models import GLBalance

logger = logging.getLogger(__name__)


class ReportPermissionChecker:
    """Rapor erişim yetki kontrolü"""

    @staticmethod
    def can_access_report(user, company, report_type):
        """Kullanıcının rapora erişim yetkisi var mı kontrol et"""
        # Süper kullanıcı her şeye erişebilir
        if user.is_superuser:
            return True

        # Şirket sahibi kontrolü
        if company.created_by == user:
            return True

        # Kullanıcının şirkete erişim yetkisi var mı?
        # TODO: CompanyUser veya benzeri bir model varsa kontrol et
        # Şimdilik sadece created_by kontrolü yapıyoruz

        # Rol bazlı kontrol
        user_groups = user.groups.values_list("name", flat=True)

        # Muhasebe raporları için gerekli roller
        report_permissions = {
            "bilanco": [
                "finance_manager",
                "accountant",
                "muhasebe_elemani",
                "kobi_owner",
            ],
            "gelir_tablosu": [
                "finance_manager",
                "accountant",
                "muhasebe_elemani",
                "kobi_owner",
            ],
            "nakit_akisi": [
                "finance_manager",
                "accountant",
                "muhasebe_elemani",
                "kobi_owner",
            ],
            "mizan": [
                "finance_manager",
                "accountant",
                "muhasebe_elemani",
                "kobi_owner",
            ],
            "yevmiye": [
                "finance_manager",
                "accountant",
                "muhasebe_elemani",
                "kobi_owner",
            ],
            "kebir": [
                "finance_manager",
                "accountant",
                "muhasebe_elemani",
                "kobi_owner",
            ],
            "kdv": ["finance_manager", "accountant", "muhasebe_elemani", "kobi_owner"],
            "muhtasar": [
                "finance_manager",
                "accountant",
                "muhasebe_elemani",
                "kobi_owner",
            ],
            "babs": ["finance_manager", "accountant", "muhasebe_elemani", "kobi_owner"],
        }

        required_roles = report_permissions.get(report_type, [])

        # Kullanıcının rolü gerekli rollerden biri mi?
        if any(role in user_groups for role in required_roles):
            return True

        return False

    @staticmethod
    def can_export_report(user, company, report_type):
        """Kullanıcının raporu dışa aktarma yetkisi var mı?"""
        if not ReportPermissionChecker.can_access_report(user, company, report_type):
            return False

        # Dışa aktarma için ek yetki kontrolü
        user_groups = user.groups.values_list("name", flat=True)
        export_roles = [
            "finance_manager",
            "accountant",
            "muhasebe_elemani",
            "kobi_owner",
        ]

        return any(role in user_groups for role in export_roles) or user.is_superuser

    @staticmethod
    def can_send_report(user, company, report_type):
        """Kullanıcının raporu gönderme yetkisi var mı?"""
        if not ReportPermissionChecker.can_export_report(user, company, report_type):
            return False

        # Gönderme için ek yetki kontrolü
        user_groups = user.groups.values_list("name", flat=True)
        send_roles = ["finance_manager", "accountant", "kobi_owner"]

        return any(role in user_groups for role in send_roles) or user.is_superuser


class KVKKDataProtection:
    """KVKK uyumlu veri koruma ve anonimleştirme"""

    @staticmethod
    def anonymize_personal_data(
        data: Dict[str, Any], fields_to_mask: List[str] = None
    ) -> Dict[str, Any]:
        """Kişisel verileri anonimleştir"""
        if fields_to_mask is None:
            fields_to_mask = [
                "tc_kimlik_no",
                "tax_number",
                "phone",
                "email",
                "address",
                "iban",
                "account_number",
                "customer_name",
                "vendor_name",
                "first_name",
                "last_name",
            ]

        anonymized = data.copy()

        for field in fields_to_mask:
            if field in anonymized:
                value = str(anonymized[field])
                if len(value) > 4:
                    # Son 4 karakteri göster, geri kalanını maskele
                    anonymized[field] = "*" * (len(value) - 4) + value[-4:]
                else:
                    anonymized[field] = "****"

        return anonymized

    @staticmethod
    def mask_sensitive_numbers(value: str, keep_last: int = 4) -> str:
        """Hassas numaraları maskele (TC, IBAN, vb.)"""
        if not value or len(value) <= keep_last:
            return "****"
        return "*" * (len(value) - keep_last) + value[-keep_last:]

    @staticmethod
    def should_anonymize_report(user, company, report_type):
        """Rapor anonimleştirilmeli mi?"""
        # Hassas raporlar için anonimleştirme gerekebilir
        sensitive_reports = ["customer_list", "vendor_list", "employee_list"]

        # Eğer kullanıcı sadece görüntüleme yetkisine sahipse anonimleştir
        user_groups = user.groups.values_list("name", flat=True)
        if "viewer" in user_groups and report_type in sensitive_reports:
            return True

        return False

    @staticmethod
    def create_data_access_log(user, company, report_type, data_fields_accessed):
        """Veri erişim logu oluştur (KVKK gereksinimi)"""
        try:
            from finance.data_security_compliance import PersonalDataRecord

            PersonalDataRecord.objects.create(
                company=company,
                data_category=None,  # TODO: Uygun kategori seç
                data_subject_id=hashlib.sha256(
                    f"{user.id}_{company.id}".encode()
                ).hexdigest()[:16],
                data_subject_type="user",
                processing_activity="ACCESS",
                processed_by=user,
                data_fields_processed=data_fields_accessed,
                processing_purpose=f"Rapor oluşturma: {report_type}",
                consent_obtained=True,  # Şirket içi raporlama için varsayılan
            )
        except Exception as e:
            logger.error(f"Veri erişim logu oluşturma hatası: {e}", exc_info=True)


class TFRSReportGenerator:
    """Türkiye Finansal Raporlama Standartları (TFRS) uyumlu rapor üretici"""

    def __init__(self, company: Company, user=None):
        self.company = company
        self.user = user
        self.currency = company.base_currency or "TRY"

    def generate_balance_sheet(
        self, as_of_date: date, comparative: bool = False
    ) -> Dict[str, Any]:
        """
        Bilanço (TFRS 1 - Finansal Tabloların Sunumu)
        Aktif ve Pasif hesaplarını TFRS formatında düzenler
        """
        # Aktif hesaplar (Varlıklar)
        asset_accounts = GLAccount.objects.filter(
            company=self.company,
            category__in=["ASSET", "CURRENT_ASSET", "NON_CURRENT_ASSET"],
            is_active=True,
        ).order_by("code")

        # Pasif hesaplar (Yükümlülükler ve Özkaynaklar)
        liability_accounts = GLAccount.objects.filter(
            company=self.company,
            category__in=["LIABILITY", "CURRENT_LIABILITY", "NON_CURRENT_LIABILITY"],
            is_active=True,
        ).order_by("code")

        equity_accounts = GLAccount.objects.filter(
            company=self.company,
            category__in=["EQUITY", "CAPITAL", "RETAINED_EARNINGS"],
            is_active=True,
        ).order_by("code")

        # Bakiye hesaplama
        def get_account_balance(account, as_of_date):
            """Hesabın belirli tarihteki bakiyesini hesapla"""
            balance = GLBalance.objects.filter(
                account=account, year=as_of_date.year, month=as_of_date.month
            ).first()

            if balance:
                return {
                    "begin_balance": balance.begin_balance or Decimal("0"),
                    "debit_total": balance.debit_total or Decimal("0"),
                    "credit_total": balance.credit_total or Decimal("0"),
                    "end_balance": (
                        (balance.begin_balance or Decimal("0"))
                        + (balance.debit_total or Decimal("0"))
                        - (balance.credit_total or Decimal("0"))
                    ),
                }

            # GLJournalLine'dan hesapla
            lines = GLJournalLine.objects.filter(
                account=account,
                entry__company=self.company,
                entry__date__lte=as_of_date,
            )

            debit_sum = lines.aggregate(Sum("debit"))["debit__sum"] or Decimal("0")
            credit_sum = lines.aggregate(Sum("credit"))["credit__sum"] or Decimal("0")

            # Aktif hesaplar için borç-alacak, pasif için alacak-borç
            if account.category in ["ASSET", "CURRENT_ASSET", "NON_CURRENT_ASSET"]:
                balance = debit_sum - credit_sum
            else:
                balance = credit_sum - debit_sum

            return {
                "begin_balance": Decimal("0"),
                "debit_total": debit_sum,
                "credit_total": credit_sum,
                "end_balance": balance,
            }

        # Aktif varlıklar
        current_assets = []
        non_current_assets = []
        total_current_assets = Decimal("0")
        total_non_current_assets = Decimal("0")

        for account in asset_accounts:
            balance_data = get_account_balance(account, as_of_date)
            balance = balance_data["end_balance"]

            if balance != 0:
                account_data = {
                    "code": account.code,
                    "name": account.name,
                    "balance": balance,
                    "category": account.category,
                }

                if account.category == "CURRENT_ASSET":
                    current_assets.append(account_data)
                    total_current_assets += balance
                else:
                    non_current_assets.append(account_data)
                    total_non_current_assets += balance

        # Pasif yükümlülükler
        current_liabilities = []
        non_current_liabilities = []
        total_current_liabilities = Decimal("0")
        total_non_current_liabilities = Decimal("0")

        for account in liability_accounts:
            balance_data = get_account_balance(account, as_of_date)
            balance = balance_data["end_balance"]

            if balance != 0:
                account_data = {
                    "code": account.code,
                    "name": account.name,
                    "balance": balance,
                    "category": account.category,
                }

                if account.category == "CURRENT_LIABILITY":
                    current_liabilities.append(account_data)
                    total_current_liabilities += balance
                else:
                    non_current_liabilities.append(account_data)
                    total_non_current_liabilities += balance

        # Özkaynaklar
        equity_items = []
        total_equity = Decimal("0")

        for account in equity_accounts:
            balance_data = get_account_balance(account, as_of_date)
            balance = balance_data["end_balance"]

            if balance != 0:
                account_data = {
                    "code": account.code,
                    "name": account.name,
                    "balance": balance,
                    "category": account.category,
                }
                equity_items.append(account_data)
                total_equity += balance

        # Kar/Zarar (Gelir Tablosu'ndan)
        income_statement = self.generate_income_statement(
            date(as_of_date.year, 1, 1), as_of_date
        )
        net_income = income_statement.get("net_income", Decimal("0"))

        # Özkaynaklara net karı ekle
        if net_income != 0:
            equity_items.append(
                {
                    "code": "590",
                    "name": "Dönem Net Karı/Zararı",
                    "balance": net_income,
                    "category": "EQUITY",
                }
            )
            total_equity += net_income

        total_assets = total_current_assets + total_non_current_assets
        total_liabilities = total_current_liabilities + total_non_current_liabilities
        total_liabilities_and_equity = total_liabilities + total_equity

        # Bilanço denkliği kontrolü
        is_balanced = abs(total_assets - total_liabilities_and_equity) < Decimal("0.01")

        return {
            "company": self.company,
            "as_of_date": as_of_date,
            "currency": self.currency,
            "current_assets": current_assets,
            "non_current_assets": non_current_assets,
            "current_liabilities": current_liabilities,
            "non_current_liabilities": non_current_liabilities,
            "equity": equity_items,
            "totals": {
                "total_current_assets": total_current_assets,
                "total_non_current_assets": total_non_current_assets,
                "total_assets": total_assets,
                "total_current_liabilities": total_current_liabilities,
                "total_non_current_liabilities": total_non_current_liabilities,
                "total_liabilities": total_liabilities,
                "total_equity": total_equity,
                "total_liabilities_and_equity": total_liabilities_and_equity,
            },
            "is_balanced": is_balanced,
            "net_income": net_income,
            "comparative": comparative,
        }

    def generate_income_statement(
        self, start_date: date, end_date: date
    ) -> Dict[str, Any]:
        """
        Gelir Tablosu (TFRS 1 - Finansal Tabloların Sunumu)
        Gelir ve gider hesaplarını TFRS formatında düzenler
        """
        # Gelir hesapları
        revenue_accounts = GLAccount.objects.filter(
            company=self.company,
            category__in=["REVENUE", "INCOME", "SALES"],
            is_active=True,
        ).order_by("code")

        # Gider hesapları (not used but kept for potential future use)
        _expense_accounts = GLAccount.objects.filter(
            company=self.company,
            category__in=["EXPENSE", "COST", "COST_OF_SALES", "OPERATING_EXPENSE"],
            is_active=True,
        ).order_by("code")

        def get_period_balance(account, start_date, end_date):
            """Hesabın dönem içindeki bakiyesini hesapla"""
            lines = GLJournalLine.objects.filter(
                account=account,
                entry__company=self.company,
                entry__date__gte=start_date,
                entry__date__lte=end_date,
            )

            debit_sum = lines.aggregate(Sum("debit"))["debit__sum"] or Decimal("0")
            credit_sum = lines.aggregate(Sum("credit"))["credit__sum"] or Decimal("0")

            # Gelir hesapları için alacak-borç, gider için borç-alacak
            if account.category in ["REVENUE", "INCOME", "SALES"]:
                return credit_sum - debit_sum
            else:
                return debit_sum - credit_sum

        # Gelirler
        revenues = []
        total_revenue = Decimal("0")

        for account in revenue_accounts:
            balance = get_period_balance(account, start_date, end_date)
            if balance != 0:
                revenues.append(
                    {"code": account.code, "name": account.name, "amount": balance}
                )
                total_revenue += balance

        # Giderler
        expenses = []

        # Satışların maliyeti
        cost_of_sales = Decimal("0")
        cost_of_sales_accounts = GLAccount.objects.filter(
            company=self.company, category="COST_OF_SALES", is_active=True
        )
        for account in cost_of_sales_accounts:
            balance = get_period_balance(account, start_date, end_date)
            cost_of_sales += balance

        # Faaliyet giderleri
        operating_expenses = Decimal("0")
        operating_expense_accounts = GLAccount.objects.filter(
            company=self.company, category="OPERATING_EXPENSE", is_active=True
        )
        for account in operating_expense_accounts:
            balance = get_period_balance(account, start_date, end_date)
            operating_expenses += balance
            expenses.append(
                {"code": account.code, "name": account.name, "amount": balance}
            )

        # Total expenses (not used but kept for potential future use)
        _total_expenses = cost_of_sales + operating_expenses

        # Brüt kar
        gross_profit = total_revenue - cost_of_sales

        # Faaliyet karı
        operating_profit = gross_profit - operating_expenses

        # Finansal gelir/gider
        financial_income = Decimal("0")
        financial_expense = Decimal("0")

        financial_accounts = GLAccount.objects.filter(
            company=self.company,
            category__in=["FINANCIAL_INCOME", "FINANCIAL_EXPENSE"],
            is_active=True,
        )
        for account in financial_accounts:
            balance = get_period_balance(account, start_date, end_date)
            if account.category == "FINANCIAL_INCOME":
                financial_income += balance
            else:
                financial_expense += balance

        # Faaliyet dışı gelir/gider
        other_income = Decimal("0")
        other_expense = Decimal("0")

        other_accounts = GLAccount.objects.filter(
            company=self.company,
            category__in=["OTHER_INCOME", "OTHER_EXPENSE"],
            is_active=True,
        )
        for account in other_accounts:
            balance = get_period_balance(account, start_date, end_date)
            if account.category == "OTHER_INCOME":
                other_income += balance
            else:
                other_expense += balance

        # Vergi öncesi kar
        profit_before_tax = (
            operating_profit
            + financial_income
            - financial_expense
            + other_income
            - other_expense
        )

        # Gelir vergisi (tahmini %20)
        income_tax_rate = Decimal("0.20")
        income_tax = (
            profit_before_tax * income_tax_rate
            if profit_before_tax > 0
            else Decimal("0")
        )

        # Net kar
        net_income = profit_before_tax - income_tax

        return {
            "company": self.company,
            "start_date": start_date,
            "end_date": end_date,
            "currency": self.currency,
            "revenues": revenues,
            "total_revenue": total_revenue,
            "cost_of_sales": cost_of_sales,
            "gross_profit": gross_profit,
            "operating_expenses": expenses,
            "total_operating_expenses": operating_expenses,
            "operating_profit": operating_profit,
            "financial_income": financial_income,
            "financial_expense": financial_expense,
            "other_income": other_income,
            "other_expense": other_expense,
            "profit_before_tax": profit_before_tax,
            "income_tax": income_tax,
            "net_income": net_income,
        }

    def generate_cash_flow_statement(
        self, start_date: date, end_date: date
    ) -> Dict[str, Any]:
        """
        Nakit Akış Tablosu (TFRS 7 - Nakit Akış Tabloları)
        Faaliyet, yatırım ve finansman faaliyetlerinden kaynaklanan nakit akışları
        """
        # Nakit hesapları
        cash_accounts = GLAccount.objects.filter(
            company=self.company,
            code__startswith="100",  # TDHP: 100 Kasa, 101 Alınan Çekler, vb.
            is_active=True,
        )

        # Banka hesapları
        bank_accounts = GLAccount.objects.filter(
            company=self.company,
            code__startswith="102",  # TDHP: 102 Bankalar
            is_active=True,
        )

        # Faaliyet faaliyetlerinden nakit akışları
        # Net kar + amortisman + işletme sermayesi değişimi
        income_statement = self.generate_income_statement(start_date, end_date)
        net_income = income_statement["net_income"]

        # Amortisman (gider hesaplarından)
        depreciation = Decimal("0")
        depreciation_accounts = GLAccount.objects.filter(
            company=self.company,
            code__startswith="77",  # TDHP: 77 Amortisman giderleri
            is_active=True,
        )
        for account in depreciation_accounts:
            lines = GLJournalLine.objects.filter(
                account=account,
                entry__company=self.company,
                entry__date__gte=start_date,
                entry__date__lte=end_date,
            )
            depreciation += lines.aggregate(Sum("debit"))["debit__sum"] or Decimal("0")

        # İşletme sermayesi değişimi
        # Alacaklar, stoklar, borçlar değişimi
        # Basitleştirilmiş hesaplama
        working_capital_change = Decimal("0")

        # Faaliyet faaliyetlerinden nakit akışı
        operating_cash_flow = net_income + depreciation + working_capital_change

        # Yatırım faaliyetlerinden nakit akışları
        # Duran varlık alımları/satışları
        investment_cash_flow = Decimal("0")
        fixed_asset_accounts = GLAccount.objects.filter(
            company=self.company,
            code__startswith="25",  # TDHP: 25 Maddi Duran Varlıklar
            is_active=True,
        )
        for account in fixed_asset_accounts:
            lines = GLJournalLine.objects.filter(
                account=account,
                entry__company=self.company,
                entry__date__gte=start_date,
                entry__date__lte=end_date,
            )
            # Alımlar (borç) - satışlar (alacak)
            investment_cash_flow += (
                lines.aggregate(Sum("debit"))["debit__sum"] or Decimal("0")
            ) - (lines.aggregate(Sum("credit"))["credit__sum"] or Decimal("0"))

        # Finansman faaliyetlerinden nakit akışları
        # Kredi alımları, özkaynak artışları, kredi ödemeleri
        financing_cash_flow = Decimal("0")
        loan_accounts = GLAccount.objects.filter(
            company=self.company,
            code__startswith="30",  # TDHP: 30 Mali Borçlar
            is_active=True,
        )
        for account in loan_accounts:
            lines = GLJournalLine.objects.filter(
                account=account,
                entry__company=self.company,
                entry__date__gte=start_date,
                entry__date__lte=end_date,
            )
            # Kredi alımları (alacak) - ödemeler (borç)
            financing_cash_flow += (
                lines.aggregate(Sum("credit"))["credit__sum"] or Decimal("0")
            ) - (lines.aggregate(Sum("debit"))["debit__sum"] or Decimal("0"))

        # Dönem başı nakit
        beginning_cash = Decimal("0")
        for account in list(cash_accounts) + list(bank_accounts):
            balance = GLBalance.objects.filter(
                account=account, year=start_date.year, month=start_date.month
            ).first()
            if balance:
                beginning_cash += balance.begin_balance or Decimal("0")

        # Net nakit artışı
        net_cash_increase = (
            operating_cash_flow + investment_cash_flow + financing_cash_flow
        )

        # Dönem sonu nakit
        ending_cash = beginning_cash + net_cash_increase

        return {
            "company": self.company,
            "start_date": start_date,
            "end_date": end_date,
            "currency": self.currency,
            "operating_activities": {
                "net_income": net_income,
                "depreciation": depreciation,
                "working_capital_change": working_capital_change,
                "operating_cash_flow": operating_cash_flow,
            },
            "investing_activities": {
                "investment_cash_flow": investment_cash_flow,
            },
            "financing_activities": {
                "financing_cash_flow": financing_cash_flow,
            },
            "net_cash_increase": net_cash_increase,
            "beginning_cash": beginning_cash,
            "ending_cash": ending_cash,
        }

    def generate_trial_balance(self, as_of_date: date) -> Dict[str, Any]:
        """
        Genel Mizan (VUK Madde 64)
        Tüm hesapların borç/alacak toplamları ve bakiyeleri
        """
        accounts = GLAccount.objects.filter(
            company=self.company, is_active=True
        ).order_by("code")

        rows = []
        total_debit = Decimal("0")
        total_credit = Decimal("0")

        for account in accounts:
            # Hesap bakiyesi
            balance = GLBalance.objects.filter(
                account=account, year=as_of_date.year, month=as_of_date.month
            ).first()

            if balance:
                debit_total = balance.debit_total or Decimal("0")
                credit_total = balance.credit_total or Decimal("0")
                end_balance = (
                    (balance.begin_balance or Decimal("0")) + debit_total - credit_total
                )
            else:
                # GLJournalLine'dan hesapla
                lines = GLJournalLine.objects.filter(
                    account=account,
                    entry__company=self.company,
                    entry__date__lte=as_of_date,
                )
                debit_total = lines.aggregate(Sum("debit"))["debit__sum"] or Decimal(
                    "0"
                )
                credit_total = lines.aggregate(Sum("credit"))["credit__sum"] or Decimal(
                    "0"
                )
                end_balance = debit_total - credit_total

            if debit_total != 0 or credit_total != 0:
                rows.append(
                    {
                        "code": account.code,
                        "name": account.name,
                        "category": account.category,
                        "begin_balance": balance.begin_balance
                        if balance
                        else Decimal("0"),
                        "debit_total": debit_total,
                        "credit_total": credit_total,
                        "end_balance": end_balance,
                    }
                )
                total_debit += debit_total
                total_credit += credit_total

        return {
            "company": self.company,
            "as_of_date": as_of_date,
            "currency": self.currency,
            "rows": rows,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "is_balanced": abs(total_debit - total_credit) < Decimal("0.01"),
        }

    def generate_journal_ledger(
        self, start_date: date, end_date: date
    ) -> Dict[str, Any]:
        """
        Yevmiye Defteri (VUK Madde 64)
        Tarih sırasına göre tüm muhasebe kayıtları
        """
        entries = (
            GLJournalEntry.objects.filter(
                company=self.company,
                date__gte=start_date,
                date__lte=end_date,
                state="posted",
            )
            .order_by("date", "number")
            .select_related("company")
        )

        rows = []
        for entry in entries:
            lines = GLJournalLine.objects.filter(entry=entry).order_by("line_no")
            for line in lines:
                rows.append(
                    {
                        "date": entry.date,
                        "entry_number": entry.number,
                        "description": entry.description,
                        "account_code": line.account.code if line.account else "",
                        "account_name": line.account.name if line.account else "",
                        "debit": line.debit or Decimal("0"),
                        "credit": line.credit or Decimal("0"),
                        "reference": entry.reference or "",
                    }
                )

        return {
            "company": self.company,
            "start_date": start_date,
            "end_date": end_date,
            "currency": self.currency,
            "rows": rows,
        }

    def generate_general_ledger(
        self, start_date: date, end_date: date
    ) -> Dict[str, Any]:
        """
        Büyük Defter (Kebir Defteri) (VUK Madde 64)
        Hesap bazında tüm hareketler
        """
        accounts = GLAccount.objects.filter(
            company=self.company, is_active=True
        ).order_by("code")

        ledger_data = []

        for account in accounts:
            lines = GLJournalLine.objects.filter(
                account=account,
                entry__company=self.company,
                entry__date__gte=start_date,
                entry__date__lte=end_date,
                entry__state="posted",
            ).order_by("entry__date", "entry__number")

            if lines.exists():
                account_lines = []
                debit_total = Decimal("0")
                credit_total = Decimal("0")

                for line in lines:
                    account_lines.append(
                        {
                            "date": line.entry.date,
                            "entry_number": line.entry.number,
                            "description": line.entry.description,
                            "debit": line.debit or Decimal("0"),
                            "credit": line.credit or Decimal("0"),
                            "balance": (line.debit or Decimal("0"))
                            - (line.credit or Decimal("0")),
                        }
                    )
                    debit_total += line.debit or Decimal("0")
                    credit_total += line.credit or Decimal("0")

                ledger_data.append(
                    {
                        "account": {
                            "code": account.code,
                            "name": account.name,
                            "category": account.category,
                        },
                        "begin_balance": Decimal(
                            "0"
                        ),  # TODO: Dönem başı bakiyesi hesapla
                        "lines": account_lines,
                        "debit_total": debit_total,
                        "credit_total": credit_total,
                        "end_balance": debit_total - credit_total,
                    }
                )

        return {
            "company": self.company,
            "start_date": start_date,
            "end_date": end_date,
            "currency": self.currency,
            "accounts": ledger_data,
        }


class ReportExportService:
    """Rapor dışa aktarma servisi (PDF, Excel, e-posta)"""

    @staticmethod
    def export_to_pdf(
        report_data: Dict[str, Any], report_type: str, template_name: str
    ) -> HttpResponse:
        """Raporu PDF olarak dışa aktar"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER
            from django.utils import timezone

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm
            )

            elements = []
            styles = getSampleStyleSheet()

            # Başlık
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=16,
                textColor=colors.HexColor("#0AAE94"),
                alignment=TA_CENTER,
                spaceAfter=30,
            )

            report_titles = {
                "bilanco": "BİLANÇO",
                "gelir_tablosu": "GELİR TABLOSU",
                "nakit_akisi": "NAKİT AKIŞ TABLOSU",
                "mizan": "GENEL MİZAN",
                "yevmiye": "YEVMİYE DEFTERİ",
                "kebir": "BÜYÜK DEFTER (KEBİR DEFTERİ)",
            }

            title = report_titles.get(report_type, "MUHASEBE RAPORU")
            elements.append(Paragraph(title, title_style))

            # Şirket bilgileri
            company = report_data.get("company")
            if company:
                company_info = f"""
                <b>{company.name}</b><br/>
                Vergi No: {company.tax_number}<br/>
                Vergi Dairesi: {company.tax_office or '-'}<br/>
                Tarih: {timezone.now().strftime('%d.%m.%Y %H:%M')}
                """
                elements.append(Paragraph(company_info, styles["Normal"]))
                elements.append(Spacer(1, 0.5 * cm))

            # Rapor içeriği
            if report_type == "bilanco":
                elements.extend(
                    ReportExportService._build_balance_sheet_table(report_data, styles)
                )
            elif report_type == "gelir_tablosu":
                elements.extend(
                    ReportExportService._build_income_statement_table(
                        report_data, styles
                    )
                )
            elif report_type == "mizan":
                elements.extend(
                    ReportExportService._build_trial_balance_table(report_data, styles)
                )
            else:
                # Genel tablo
                elements.append(Paragraph("Rapor detayları", styles["Heading2"]))

            # KVKK uyarısı
            kvkk_notice = """
            <i><font size="8">
            Bu rapor KVKK (6698 sayılı Kanun) kapsamında kişisel veri içerebilir. 
            Raporun gizliliği ve güvenliği raporu alan kişinin sorumluluğundadır.
            </font></i>
            """
            elements.append(Spacer(1, 1 * cm))
            elements.append(Paragraph(kvkk_notice, styles["Normal"]))

            # PDF oluştur
            doc.build(elements)
            buffer.seek(0)

            response = HttpResponse(buffer, content_type="application/pdf")
            filename = f"{report_type}_{company.slug if company else 'report'}_{timezone.now().strftime('%Y%m%d')}.pdf"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            logger.error(f"PDF oluşturma hatası: {e}", exc_info=True)
            return HttpResponse(f"PDF oluşturma hatası: {str(e)}", status=500)

    @staticmethod
    def _build_balance_sheet_table(report_data: Dict[str, Any], styles) -> List:
        """Bilanço tablosu oluştur"""
        from reportlab.platypus import Table, TableStyle, Spacer
        from reportlab.lib import colors
        from reportlab.lib.units import cm

        elements = []

        # Aktif varlıklar
        data = [["Hesap Kodu", "Hesap Adı", "Tutar"]]

        current_assets = report_data.get("current_assets", [])
        for asset in current_assets:
            data.append([asset["code"], asset["name"], f"{asset['balance']:,.2f}"])

        if data:
            table = Table(data, colWidths=[3 * cm, 8 * cm, 3 * cm])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )
            elements.append(table)
            elements.append(Spacer(1, 0.5 * cm))

        return elements

    @staticmethod
    def _build_income_statement_table(report_data: Dict[str, Any], styles) -> List:
        """Gelir tablosu oluştur"""
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm

        elements = []

        data = [["Kalem", "Tutar"]]

        # Gelirler
        revenues = report_data.get("revenues", [])
        for revenue in revenues:
            data.append([revenue["name"], f"{revenue['amount']:,.2f}"])

        data.append(["TOPLAM GELİR", f"{report_data.get('total_revenue', 0):,.2f}"])
        data.append(["", ""])  # Boş satır

        # Giderler
        expenses = report_data.get("expenses", [])
        for expense in expenses:
            data.append([expense["name"], f"{expense['amount']:,.2f}"])

        data.append(["TOPLAM GİDER", f"{report_data.get('total_expenses', 0):,.2f}"])
        data.append(["NET KAR", f"{report_data.get('net_income', 0):,.2f}"])

        table = Table(data, colWidths=[10 * cm, 4 * cm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(table)

        return elements

    @staticmethod
    def _build_trial_balance_table(report_data: Dict[str, Any], styles) -> List:
        """Mizan tablosu oluştur"""
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm

        elements = []

        data = [["Hesap Kodu", "Hesap Adı", "Borç", "Alacak", "Bakiye"]]

        rows = report_data.get("rows", [])
        for row in rows[:50]:  # İlk 50 satır (sayfa sınırı)
            data.append(
                [
                    row["code"],
                    row["name"],
                    f"{row['debit_total']:,.2f}",
                    f"{row['credit_total']:,.2f}",
                    f"{row['end_balance']:,.2f}",
                ]
            )

        if len(rows) > 50:
            data.append(["...", f"Toplam {len(rows)} hesap", "", "", ""])

        totals = report_data.get("totals", {})
        data.append(
            [
                "TOPLAM",
                "",
                f"{totals.get('total_debit', 0):,.2f}",
                f"{totals.get('total_credit', 0):,.2f}",
                "",
            ]
        )

        table = Table(data, colWidths=[2 * cm, 6 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(table)

        return elements

    @staticmethod
    def export_to_excel(report_data: Dict[str, Any], report_type: str) -> HttpResponse:
        """Raporu Excel olarak dışa aktar"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from django.utils import timezone

            output = io.BytesIO()
            wb = Workbook()
            ws = wb.active

            # Başlık
            report_titles = {
                "bilanco": "BİLANÇO",
                "gelir_tablosu": "GELİR TABLOSU",
                "nakit_akisi": "NAKİT AKIŞ TABLOSU",
                "mizan": "GENEL MİZAN",
                "yevmiye": "YEVMİYE DEFTERİ",
                "kebir": "BÜYÜK DEFTER",
            }

            title = report_titles.get(report_type, "MUHASEBE RAPORU")
            ws["A1"] = title
            ws["A1"].font = Font(size=16, bold=True, color="FF0AAE94")
            ws.merge_cells("A1:E1")

            # Şirket bilgileri
            company = report_data.get("company")
            if company:
                ws["A3"] = f"Şirket: {company.name}"
                ws["A4"] = f"Vergi No: {company.tax_number}"
                ws["A5"] = f"Tarih: {timezone.now().strftime('%d.%m.%Y %H:%M')}"

            # Rapor verilerini Excel'e yaz
            row = 7
            if report_type == "mizan":
                # Mizan için
                headers = ["Hesap Kodu", "Hesap Adı", "Borç", "Alacak", "Bakiye"]
                ws.append(headers)

                # Başlık formatı
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid"
                    )

                rows = report_data.get("rows", [])
                for data_row in rows:
                    row += 1
                    ws.append(
                        [
                            data_row["code"],
                            data_row["name"],
                            float(data_row["debit_total"]),
                            float(data_row["credit_total"]),
                            float(data_row["end_balance"]),
                        ]
                    )

            # Sütun genişliklerini ayarla
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (AttributeError, TypeError, ValueError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            filename = f"{report_type}_{company.slug if company else 'report'}_{timezone.now().strftime('%Y%m%d')}.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            logger.error(f"Excel oluşturma hatası: {e}", exc_info=True)
            return HttpResponse(f"Excel oluşturma hatası: {str(e)}", status=500)

    @staticmethod
    def send_report_email(
        user,
        company: Company,
        report_data: Dict[str, Any],
        report_type: str,
        recipients: List[str],
        pdf_file: Optional[io.BytesIO] = None,
    ) -> bool:
        """Raporu e-posta ile gönder"""
        try:
            report_titles = {
                "bilanco": "Bilanço",
                "gelir_tablosu": "Gelir Tablosu",
                "nakit_akisi": "Nakit Akış Tablosu",
                "mizan": "Genel Mizan",
                "yevmiye": "Yevmiye Defteri",
                "kebir": "Büyük Defter",
            }

            title = report_titles.get(report_type, "Muhasebe Raporu")

            context = {
                "user": user,
                "company": company,
                "report_type": title,
                "report_data": report_data,
                "generated_at": timezone.now(),
                "site_url": getattr(settings, "SITE_URL", "https://finasis.com.tr"),
            }

            # HTML e-posta içeriği
            html_content = render_to_string(
                "accounting/emails/report_email.html", context
            )
            text_content = strip_tags(html_content)

            # E-posta gönder
            email = EmailMultiAlternatives(
                subject=f"{title} - {company.name}",
                body=text_content,
                from_email=getattr(
                    settings, "DEFAULT_FROM_EMAIL", "noreply@finasis.com.tr"
                ),
                to=recipients,
            )
            email.attach_alternative(html_content, "text/html")

            # PDF ekle (varsa)
            if pdf_file:
                pdf_file.seek(0)
                filename = f"{report_type}_{company.slug}_{timezone.now().strftime('%Y%m%d')}.pdf"
                email.attach(filename, pdf_file.read(), "application/pdf")

            email.send()

            # KVKK: E-posta gönderim logu
            KVKKDataProtection.create_data_access_log(
                user, company, report_type, ["email_sent", "recipients_count"]
            )

            logger.info(
                f"Rapor e-postası gönderildi: {report_type}, "
                f"Company {company.id}, Recipients {len(recipients)}"
            )

            return True

        except Exception as e:
            logger.error(f"Rapor e-posta gönderme hatası: {e}", exc_info=True)
            return False
